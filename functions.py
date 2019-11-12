import openpyxl
from openpyxl import load_workbook
from PIL import Image, ImageTk
import time

#variables for Excel
wb = openpyxl.load_workbook('North Spine Canteen Details (New).xlsx') #load excel
openinghours = wb['opening hours']
food = wb['food']
info = wb['store information']
storeinfo = {}

#Yi Shen
def excelConversion(wb):                                         #Parses through the excel and generates 2 dictionaries: menu, storeOpen
    openinghours = wb['opening hours']                                          #opening hours sheet
    food = wb['food']                                                           #food sheet
    info = wb['store information']                                         #store information sheet
    menu = {}
    storeOpen = {}
    
    for row in range(2, food.max_row):                                          #parses information in food sheet and converts to dictionary
        key = (food.cell(row, 1).value, food.cell(row, 2).value)
        menu.update({key : (food.cell(row, 3).value, food.cell(row, 4).value)})
    
    for row in range(2, info.max_row):
        storeinfo.update({info.cell(row, 1).value : (info.cell(row, 2).value, info.cell(row, 3).value, info.cell(row, 4).value, info.cell(row, 5).value)})

    DayDict = {}
    PreviousStore = openinghours.cell(2, 1).value

    for row in range(2, openinghours.max_row):                                   #iterates through the rows of opening hours sheet
        for row in range(2, openinghours.max_row):                                   #iterates through the rows of opening hours sheet
            if (openinghours.cell(row, 1).value != PreviousStore):                   #if the name of the store changes, update dictionary with store name and the days that the store is open
                storeOpen.update({PreviousStore : DayDict})
                PreviousStore = openinghours.cell(row, 1).value
                DayDict = {}                                                         #Clear dictionary

            if ((openinghours.cell(row, 2).value) == 'Weekdays'):           #Update DayDict with key as the day of the week(0 = Monday ... 6 = Sunday) and the value as a tuple with (opening hours, closing hours)
                for n in range(0, 5):
                    DayDict.update({n : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Weekends'):
                for n in range(5, 7):
                    DayDict.update({n : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Monday'):
                DayDict.update({0 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Tuesday'):
                DayDict.update({1 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Wednesday'):
                DayDict.update({2 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Thursday'):
                DayDict.update({3 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Friday'):
                DayDict.update({4 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Saturday'):
                DayDict.update({5 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Sunday'):
                DayDict.update({6 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

    return menu, storeOpen      #format of menu dictionary --> {(Store Name, Food Name) : Price} | format of storeOpen dictionary --> {Store Name : {Day of week, (Opening Time, Closing Time)}}
