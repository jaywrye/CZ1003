import menuDB
import os
from datetime import *
from openpyxl import load_workbook

def excelConversion(xlsxFileLocation):                                         #Parses through the excel and generates 2 dictionaries: menu, storeOpen
    wb = load_workbook(xlsxFileLocation)

    openinghours = wb['opening hours']                                          #opening hours sheet
    food = wb['food']                                                           #food sheet
    storeinfo = wb['store information']                                         #store information sheet
    menu = {}
    storeOpen = {}

    for row in range(2, food.max_row):                                          #parses information in food sheet and converts to dictionary
        key = (food.cell(row, 1).value, food.cell(row, 2).value)
        menu.update({key : (food.cell(row, 3).value, food.cell(row, 4).value)})

    DayDict = {}
    PreviousStore = openinghours.cell(2, 1).value

    for row in range(2, openinghours.max_row):                                   #iterates through the rows of opening hours sheet
        if (openinghours.cell(row, 1).value != PreviousStore):                   #if the name of the store changes, update dictionary with store name and the days that the store is open
            storeOpen.update({PreviousStore : DayDict})
            PreviousStore = openinghours.cell(row, 1).value
            DayDict = {}                                                         #Clear dictionary

        if ((openinghours.cell(row, 2).value) == 'Weekdays'):           #Update DayDict with key as the day of the week(0 = Monday ... 6 = Sunday) and the value as a tuple with (opening hours, closing hours)
            for n in range(0, 5):
                DayDict.update({n : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

        if ((openinghours.cell(row, 2).value) == 'Weekends'):
            for n in range(5, 7):
                DayDict.update({n : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

        if ((openinghours.cell(row, 2).value) == 'Monday'):
            DayDict.update({0 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

        if ((openinghours.cell(row, 2).value) == 'Tuesday'):
            DayDict.update({1 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

        if ((openinghours.cell(row, 2).value) == 'Wednesday'):
            DayDict.update({2 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

        if ((openinghours.cell(row, 2).value) == 'Thursday'):
            DayDict.update({3 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

        if ((openinghours.cell(row, 2).value) == 'Friday'):
            DayDict.update({4 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

        if ((openinghours.cell(row, 2).value) == 'Saturday'):
            DayDict.update({5 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

        if ((openinghours.cell(row, 2).value) == 'Sunday'):
            DayDict.update({6 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

    return menu, storeOpen

def excelConversionInfo(xlsxFileLocation):                                         #Parses through the excel and generates 1 dictionaries: storeinfo
    wb = load_workbook(xlsxFileLocation)
    info = wb['store information']
    storeinfo = {}

    for row in range(2, info.max_row):
        storeinfo.update({info.cell(row, 1).value : (info.cell(row, 2).value, info.cell(row, 3).value, info.cell(row, 4).value, info.cell(row, 5).value)})
    
    return storeinfo

def storesopen(dayofweek, storeOpenDict):                                          # Function that returns a list of stores open on the chosen day using chosen Dict
    storesopenList = []                                                        # dayofweek is an int ranging from 0-6  with 0 being Monday and 6 being Sunday
    for n in storeOpenDict:
        if dayofweek in storeOpenDict[n]:
            storesopenList.append(n)
    return storesopenList

def storeopeningtime(storename, day, storeOpenDict):                             # Function that returns opening time and closing time based on storename and day
    if day in storeOpenDict[storename]:
        opentime = storeOpenDict[storename][day][0]
        closetime = storeOpenDict[storename][day][1]
    else:
        day = next(iter(storeOpenDict[storename]))
        opentime = storeOpenDict[storename][day][0]
        closetime = storeOpenDict[storename][day][1]
    return opentime, closetime

def storemenu(storename, menuDict):                                            # Function that returns dictionary of ALL items on their menu {"Name of food item" : price of food}
    storemenuDict = {}                                                         # storename is a string
    for n in menuDict:
        if storename in n:
            storemenuDict.update({n[1] : menuDict[n][0]})
    return storemenuDict

def storemenutime(storename, currenttime, day, menuDict, storeOpenDict):         # Function that returns dictionary of ALL items on their menu available at current time {"Name of food item" : price of food}
    storemenuDict = {}                                                           # storename is a string
    lunchtime = time(12, 0, 0)
    openingtime, closingtime = storeopeningtime(storename, day, storeOpenDict)
    for n in menuDict:
        if storename in n:
            availability = menuDict[n][1]
            if availability == "All" and currenttime >= openingtime and currenttime <= closingtime:
                storemenuDict.update({n[1] : menuDict[n][0]})
            if availability == "Breakfast" and currenttime <= lunchtime and currenttime >= openingtime:
                storemenuDict.update({n[1] : menuDict[n][0]})
            if availability == "Lunch" and currenttime >= lunchtime and currenttime <= closingtime:
                storemenuDict.update({n[1] : menuDict[n][0]})
    return storemenuDict

def searchfood(foodname, menuDict):                                                   # Function that returns price for chosen food
    for n in menuDict:
        if n[1] == foodname:
            return menuDict[n][0]

def dictSplit(storemenuDict):                                   # Function that splits a dictionary into 2 lists, one that has food name another that has price
    storenameList = []
    storepriceList = []
    for storename in storemenuDict:
        storenameList.append(storename)
        storepriceList.append(storemenuDict[storename])
    return storenameList, storepriceList

def storeAvgWaitTime(storename, customerNumber, infoDict):                                #Function that returns the average wait time per person for that store
    storetype = ""
    avgwaittime = 0
    for n in infoDict:
        if storename in infoDict:
            infoDict = infoDict[n][0]
        return int(avgwaittime)*customerNumber


