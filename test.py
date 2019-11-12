import tkinter as tk
import tkinter.ttk as ttk
from tkinter import *   
import datetime
import time
from tkinter import messagebox

import os
import openpyxl
from openpyxl import load_workbook
from PIL import Image, ImageTk
from tkcalendar import Calendar, DateEntry

#basic layout of page
root = tk.Tk() #initialise window
root.title("Home") #create window
root.config(bg = 'white') #change background color
root.minsize(800, 500) #set size of window 


#declarations
day = datetime.datetime.today().weekday()
storemenuDict = {}
storesOpenList = []

#declare variables
stall = ""
opentime = 0
closetime = 0
time1 = ''
userInputFood = StringVar()
pplwaiting = StringVar()
wb = openpyxl.load_workbook('North Spine Canteen Details (New).xlsx') #load excel

openinghours = wb['opening hours']
food = wb['food']
info = wb['store information']
menu = {}
menuDict = {}
storeOpen = {}
storeinfo = {}
storeOpenDict = {}
DayDict = {}
dayofweek = datetime.datetime.today().weekday()

#Yi Shen
def excelConversion(wb):                                         #Parses through the excel and generates 2 dictionaries: menu, storeOpen
    openinghours = wb['opening hours']                                          #opening hours sheet
    food = wb['food']                                                           #food sheet
    info = wb['store information']                                         #store information sheet
    menu = {}
    storeOpen = {}
    
    for row in range(2, food.max_row):                                          #parses information in food sheet and converts to dictionary
        key = (food.cell(row, 1).value, food.cell(row, 2).value)
        menu.update({key : (food.cell(row, 3).value, food.cell(row, 4).value)})
    
    for row in range(2, info.max_row):
        storeinfo.update({info.cell(row, 1).value : (info.cell(row, 2).value, info.cell(row, 3).value, info.cell(row, 4).value, info.cell(row, 5).value)})

    DayDict = {}
    PreviousStore = openinghours.cell(2, 1).value

    for row in range(2, openinghours.max_row):                                   #iterates through the rows of opening hours sheet
        for row in range(2, openinghours.max_row):                                   #iterates through the rows of opening hours sheet
            if (openinghours.cell(row, 1).value != PreviousStore):                   #if the name of the store changes, update dictionary with store name and the days that the store is open
                storeOpen.update({PreviousStore : DayDict})
                PreviousStore = openinghours.cell(row, 1).value
                DayDict = {}                                                         #Clear dictionary

            if ((openinghours.cell(row, 2).value) == 'Weekdays'):           #Update DayDict with key as the day of the week(0 = Monday ... 6 = Sunday) and the value as a tuple with (opening hours, closing hours)
                for n in range(0, 5):
                    DayDict.update({n : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Weekends'):
                for n in range(5, 7):
                    DayDict.update({n : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Monday'):
                DayDict.update({0 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Tuesday'):
                DayDict.update({1 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Wednesday'):
                DayDict.update({2 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Thursday'):
                DayDict.update({3 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Friday'):
                DayDict.update({4 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Saturday'):
                DayDict.update({5 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

            if ((openinghours.cell(row, 2).value) == 'Sunday'):
                DayDict.update({6 : (openinghours.cell(row, 3).value, openinghours.cell(row, 4).value)})

    return menu, storeOpen      #format of menu dictionary --> {(Store Name, Food Name) : Price} | format of storeOpen dictionary --> {Store Name : {Day of week, (Opening Time, Closing Time)}}

def userWaitingTime():
    global pplwaiting
    try:
        customernumber = int(pplwaiting.get())
        avgwaittime = int(storeinfo[store][3])
        waitingtime = customernumber * avgwaittime
        messagebox.showinfo("Waiting Time", "The waiting time will be approximately " + str(waitingtime) + " mins")
    except:
        messagebox.showinfo("Error", "Please enter numbers only.")

def usercustom():
    storesOpenList = []
    global ddl, convertedtime2

    # get user input date
    userdate = cal.get_date()
    
    # get user input time
    userhour = hourcombo.get()
    if userhour.isdigit():
            userhour = hourcombo.get()
    else:
        messagebox.showinfo("Error", "Please enter numbers only.")

    usermin = mincombo.get()
    if usermin.isdigit():
            usermin = hourcombo.get()
    else:
        messagebox.showinfo("Error", "Please enter numbers only.")

    usertime = userhour + ":" + usermin
    convertedtime = datetime.datetime.strptime(usertime, "%H:%M")
    convertedtime2 = convertedtime.time()

    #returns list of open stalls based on user defined date and time
    for n in storeOpenDict:
        if dayofweek in storeOpenDict[n]:
            try: 
                if(convertedtime2 >= storeOpenDict[n][dayofweek][0]) and (convertedtime2 <= storeOpenDict[n][dayofweek][1]):
                    storesOpenList.append(n)
                    print(storesOpenList)
            except:
                break

    ddl = StringVar(window2)
    try:
        ddl.set(storesOpenList[0]) #set first item as default value
        ddlLabel = Label(window2, text = "Please choose a restaurant: ", font = h3, bg = "white")
        ddlLabel.place(x = 10, y = 110)
        ddlist = OptionMenu(window2, ddl, storesOpenList[0], *storesOpenList)
        ddlist.config(bg = "white", width = 20, relief = FLAT, font = h4)
        ddlist.place(x = 250, y = 110)

        #gobtn
        load = Image.open('images/gobtn.png')
        render = ImageTk.PhotoImage(load)
        img = Button(window2, image = render, relief = FLAT, borderwidth = 0, command = customshopmenu)
        img.image = render
        img.place(x = 450, y = 107) 
    except:
        messagebox.showerror("Error", "Sorry, no stores are available at your chosen time/date.")
        

def customshopmenu():
    global fooditems, price
    chosenshop = ddl.get()
    titleLabel = ""

    titleLabel = Label(window2, text = chosenshop + "'s Menu", bg = "white", font = h2)
    titleLabel.place(x = 10, y = 150)

    storemenutime(chosenshop, convertedtime2, day, menuDict, storeOpenDict)

    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(window2, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 10, y = 200)
    price = Label(window2, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 350, y = 200)

def allstores():
    window = tk.Toplevel(root)
    window.config(bg = "white")
    window.title("All Stores")
    window.minsize(600, 450) #set size of window 

    titlelabel = Label(window, text = "All Stores", font = h1, bg = "white")
    titlelabel.place(x = 255, y = 20)

    load = Image.open('images/miniwokbtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = miniwokbtn)
    img.image = render
    img.place(x = 30, y = 90)

    load = Image.open('images/ytfbtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = yongtaufoobtn)
    img.image = render
    img.place(x = 210, y = 90)

    load = Image.open('images/chickenricebtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = chickricebtn)
    img.image = render
    img.place(x = 390, y = 90)

    load = Image.open('images/noodlesbtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = noodlesbtn)
    img.image = render
    img.place(x = 30, y = 150)

    load = Image.open('images/mixedricebtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = westernbtn)
    img.image = render
    img.place(x = 210, y = 150)

    load = Image.open('images/westernbtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = westernbtn)
    img.image = render
    img.place(x = 390, y = 150)

    load = Image.open('images/drinksbtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = drinksbtn)
    img.image = render
    img.place(x = 30, y = 210)

    load = Image.open('images/soupbtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = soupbtn)
    img.image = render
    img.place(x = 210, y = 210)

    load = Image.open('images/malaybtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = malaybtn)
    img.image = render
    img.place(x = 390, y = 210)

    load = Image.open('images/vegetarianbtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = vegetarianbtn)
    img.image = render
    img.place(x = 30, y = 270)

    load = Image.open('images/saladbtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = saladbtn)
    img.image = render
    img.place(x = 210, y = 270)

    load = Image.open('images/pastabtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = pastabtn)
    img.image = render
    img.place(x = 390, y = 270)

    load = Image.open('images/ljsbtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = longjohnbtn)
    img.image = render
    img.place(x = 30, y = 330)

    load = Image.open('images/macbtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window, image = render, borderwidth = 0, relief = FLAT, command = macbtn)
    img.image = render
    img.place(x = 210, y = 330)

def custom():
    global window2, cal, hourcombo, mincombo
    window2 = tk.Toplevel(root)
    window2.config(bg = "white")
    window2.title("Customisation")
    window2.minsize(600, 500) #set size of window 

    Label(window2, text = "Customisation", font = h1, bg = "white").place(x = 10, y = 20)
    style = ttk.Style()
    style.configure("TButton", font = ("Century Gothic", 12, 'bold'), foreground = "#3366CC", background = 'white', width = "15")
    #calendar
    Label(window2, text = 'Choose date', font = h3, bg = "white").place(x = 10, y = 60)
    cal = DateEntry(window2, background='darkblue', foreground='white')
    cal.place(x = 130, y = 65)

    #time
    Label(window2, text = "Choose time", font = h3, bg = "white").place(x = 240, y = 60)

    hour_list = []
    for hour in range(0, 24):
        if hour < 10:
            hour = '0' + str(hour)
        hour_list.append(str(hour))
    hourcombo = ttk.Combobox(window2, values=hour_list, width=5,)
    hourcombo.set(10)
    hourcombo.place(x = 350, y = 65)

    minute_list = []
    for minute in range(0,60):
        if minute < 10:
            minute = '0' + str(minute)
        minute_list.append(str(minute))
    mincombo = ttk.Combobox(window2, text = "00", values=minute_list, width=5)
    mincombo.set(00)
    mincombo.place(x = 420, y = 65)

    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(window2, image = render, relief = FLAT, borderwidth = 0, command = usercustom)
    img.image = render
    img.place(x = 480, y = 58)
    

def storeopeningtime(store, day, storeOpenDict):                             # Function that returns opening time and closing time based on storename and day
    global opentime, closetime
    if day in storeOpenDict[store]:
        opentime = storeOpenDict[store][day][0]
        closetime = storeOpenDict[store][day][1]
    else:
        day = next(iter(storeOpenDict[store]))
        opentime = storeOpenDict[store][day][0]
        closetime = storeOpenDict[store][day][1]

def storemenutime(store, current, day, menuDict, storeOpenDict):         # Function that returns dictionary of ALL items on their menu available at current time {"Name of food item" : price of food}
    global storemenuDict
    storemenuDict = {}                                                         # storename is a string
    lunchtime = datetime.time(12,0,0)
    storeopeningtime(store, day, storeOpenDict)
    for n in menuDict:
        if store in n:
            availability = menuDict[n][1]
            if availability == "All":
                storemenuDict.update({n[1] : menuDict[n][0]})
            if availability == "Breakfast" and current <= lunchtime and current >= opentime:
                storemenuDict.update({n[1] : menuDict[n][0]})
            if availability == "Lunch" and current >= lunchtime and current <= closetime:
                storemenuDict.update({n[1] : menuDict[n][0]})

def miniwokbtn():
    global store 
    store = "Mini Wok"

    storeopeningtime(store, day, storeOpenDict)

    miniwokpop = tk.Toplevel(root)
    miniwokpop.minsize(450, 700)
    miniwokpop.config(bg = "white")
    miniwokpop.title(store)

    #banner image
    load = Image.open('images/miniwok.png')
    render = ImageTk.PhotoImage(load)
    img = Label(miniwokpop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(miniwokpop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(miniwokpop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(miniwokpop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(miniwokpop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(miniwokpop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(miniwokpop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)

    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(miniwokpop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 30, y = 290)
    price = Label(miniwokpop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 350, y = 290)

def yongtaufoobtn():
    global store
    store = "Yong Tau Foo"

    storeopeningtime(store, day, storeOpenDict)

    yongtaufoopop = tk.Toplevel(root)
    yongtaufoopop.minsize(450, 580)
    yongtaufoopop.config(bg = "white")
    yongtaufoopop.title(store)

    load = Image.open('images/ytf.png')
    render = ImageTk.PhotoImage(load)
    img = Label(yongtaufoopop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(yongtaufoopop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(yongtaufoopop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(yongtaufoopop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(yongtaufoopop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(yongtaufoopop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(yongtaufoopop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)

    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(yongtaufoopop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 50, y = 300)
    price = Label(yongtaufoopop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 340, y = 300)

def chickricebtn():
    global store
    store = "Chicken Rice"

    storeopeningtime(store, day, storeOpenDict)

    chickricepop = tk.Toplevel(root)
    chickricepop.minsize(450, 520)
    chickricepop.config(bg = "white")
    chickricepop.title(store)

    load = Image.open('images/chickenrice.png')
    render = ImageTk.PhotoImage(load)
    img = Label(chickricepop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(chickricepop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(chickricepop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(chickricepop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(chickricepop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(chickricepop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(chickricepop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 260)

    storemenutime(store, current, day, menuDict, storeOpenDict)

    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(chickricepop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 50, y = 300)
    price = Label(chickricepop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 340, y = 300)

def noodlesbtn():
    global store
    store = "Noodles"

    storeopeningtime(store, day, storeOpenDict)

    noodlespop = tk.Toplevel(root)
    noodlespop.minsize(450, 600)
    noodlespop.config(bg = "white")
    noodlespop.title(store)

    load = Image.open('images/noodles.png')
    render = ImageTk.PhotoImage(load)
    img = Label(noodlespop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(noodlespop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 185)
    ophours = Label(noodlespop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 185)

    Label(noodlespop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(noodlespop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(noodlespop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(noodlespop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)

    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(noodlespop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 50, y = 300)
    price = Label(noodlespop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 340, y = 300)

def mixedricebtn():
    global store
    store = "Mixed Rice"

    storeopeningtime(store, day, storeOpenDict)

    mixedricepop = tk.Toplevel(root)
    mixedricepop.minsize(450, 460)
    mixedricepop.config(bg = "white")
    mixedricepop.title(store)

    load = Image.open('images/mixedrice.png')
    render = ImageTk.PhotoImage(load)
    img = Label(mixedricepop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(mixedricepop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(mixedricepop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(mixedricepop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(mixedricepop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(mixedricepop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(mixedricepop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)

    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(mixedricepop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 100, y = 300)
    price = Label(mixedricepop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 250, y = 300)

def westernbtn():
    global store
    store = "Western"

    storeopeningtime(store, day, storeOpenDict)

    westernpop = tk.Toplevel(root)
    westernpop.minsize(450, 500)
    westernpop.config(bg = "white")
    westernpop.title(store)

    load = Image.open('images/western.png')
    render = ImageTk.PhotoImage(load)
    img = Label(westernpop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(westernpop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(westernpop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(westernpop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(westernpop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(westernpop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(westernpop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)

    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(westernpop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 50, y = 300)
    price = Label(westernpop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 340, y = 300)

def drinksbtn():
    global store
    store = "Drinks"

    storeopeningtime(store, day, storeOpenDict)

    drinkspop = tk.Toplevel(root)
    drinkspop.minsize(450, 700)
    drinkspop.title(store)
    drinkspop.config(bg = "white")

    load = Image.open('images/drinks.png')
    render = ImageTk.PhotoImage(load)
    img = Label(drinkspop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(drinkspop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(drinkspop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(drinkspop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(drinkspop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(drinkspop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(drinkspop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)
        
    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(drinkspop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 50, y = 300)
    price = Label(drinkspop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 340, y = 300)

def soupbtn():
    global store
    store = "Soup Delight"

    storeopeningtime(store, day, storeOpenDict)

    souppop = tk.Toplevel(root)
    souppop.minsize(450, 560)
    souppop.title(store)
    souppop.config(bg = "white")

    load = Image.open('images/soup.png')
    render = ImageTk.PhotoImage(load)
    img = Label(souppop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(souppop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(souppop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(souppop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(souppop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(souppop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(souppop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)

    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(souppop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 50, y = 300)
    price = Label(souppop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 300, y = 300)

def malaybtn():
    global store
    store = "Malay BBQ"

    storeopeningtime(store, day, storeOpenDict)

    malaypop = tk.Toplevel(root)
    malaypop.minsize(450, 530)
    malaypop.title(store)
    malaypop.config(bg = "white")

    load = Image.open('images/malay.png')
    render = ImageTk.PhotoImage(load)
    img = Label(malaypop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(malaypop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(malaypop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(malaypop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(malaypop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(malaypop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(malaypop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)
        
    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(malaypop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 50, y = 300)
    price = Label(malaypop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 300, y = 300)

def vegetarianbtn():
    global store
    store = "Vegetarian"

    storeopeningtime(store, day, storeOpenDict)

    vegetarianpop = tk.Toplevel(root)
    vegetarianpop.minsize(450, 500)
    vegetarianpop.title(store)
    vegetarianpop.config(bg = "white")

    load = Image.open('images/vegetarian.png')
    render = ImageTk.PhotoImage(load)
    img = Label(vegetarianpop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(vegetarianpop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(vegetarianpop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(vegetarianpop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(vegetarianpop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(vegetarianpop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(vegetarianpop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)
    
    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(vegetarianpop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 50, y = 300)
    price = Label(vegetarianpop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 340, y = 300)

def saladbtn():
    global store
    store = "Salad"

    storeopeningtime(store, day, storeOpenDict)

    saladpop = tk.Toplevel(root)
    saladpop.minsize(450, 580)
    saladpop.title(store)
    saladpop.config(bg = "white")

    load = Image.open('images/salad.png')
    render = ImageTk.PhotoImage(load)
    img = Label(saladpop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(saladpop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(saladpop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(saladpop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(saladpop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(saladpop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(saladpop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)

    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(saladpop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 50, y = 300)
    price = Label(saladpop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 300, y = 300)

def pastabtn():
    global store
    store = "Pasta"

    storeopeningtime(store, day, storeOpenDict)

    pastapop = tk.Toplevel(root)
    pastapop.minsize(450, 700)
    pastapop.config(bg = "white")
    pastapop.title(store)

    load = Image.open('images/pasta.png')
    render = ImageTk.PhotoImage(load)
    img = Label(pastapop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(pastapop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(pastapop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(pastapop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(pastapop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(pastapop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(pastapop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)

    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(pastapop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 50, y = 300)
    price = Label(pastapop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 340, y = 300)

def longjohnbtn():
    global store
    store = "Long John Silver"

    storeopeningtime(store, day, storeOpenDict)

    ljspop = tk.Toplevel(root)
    ljspop.minsize(450, 650)
    ljspop.title(store)
    ljspop.config(bg = "white")

    load = Image.open('images/ljs.png')
    render = ImageTk.PhotoImage(load)
    img = Label(ljspop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(ljspop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(ljspop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(ljspop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(ljspop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(ljspop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(ljspop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)
    
    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(ljspop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 30, y = 300)
    price = Label(ljspop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 370, y = 300)
    
def macbtn():
    global store
    store = "McDonalds"

    storeopeningtime(store, day, storeOpenDict)

    macpop = tk.Toplevel(root)
    macpop.minsize(450, 450)
    macpop.title(store)
    macpop.config(bg = "white")

    load = Image.open('images/mcdonalds.png')
    render = ImageTk.PhotoImage(load)
    img = Label(macpop, image = render)
    img.image = render
    img.place(x = 0, y = 1)

    oplabel = Label(macpop, text = "Operating hours: ", font = h3Bold, bg = "white")
    oplabel.place(x = 10, y = 180)
    ophours = Label(macpop, text = (str(opentime) + " to " + str(closetime)), font = h3, bg = "white")
    ophours.place(x = 150, y = 180)

    Label(macpop, text = "No. of people: ", font = h3, bg = "white").place(x = 10, y = 215)
    Entry(macpop, textvariable = pplwaiting).place(x = 130, y = 220)

    #go button
    load = Image.open('images/gobtn.png')
    render = ImageTk.PhotoImage(load)
    img = Button(macpop, image = render, relief = FLAT, borderwidth = 0, command = userWaitingTime)
    img.image = render
    img.place(x = 270, y = 215)

    menulabel = Label(macpop, text = "Menu", font = h2, bg = "white")
    menulabel.place(x = 200, y = 270)

    storemenutime(store, current, day, menuDict, storeOpenDict)
    
    allprices = [v for v in storemenuDict.values()]
    fooditems = Label(macpop, text = '\n'.join(storemenuDict), font = h3, bg = "white", justify = LEFT)
    fooditems.place(x = 50, y = 300)
    price = Label(macpop, text = '\n'.join(map(str, allprices)), font = h3, bg = "white")
    price.place(x = 340, y = 300)

def tick():
    global time1
    time2 = time.strftime('%H:%M:%S')

    if time2 != time1:
        time1 = time2
        clock.config(text = time2)

    clock.after(10, tick)

# Font Styling
h1 = ("Century Gothic", 14, "bold")
h2 =  ("Century Gothic", 12, "bold")
h3Italic = ("Century Gothic", 11, "italic")
h3 =  ("Century Gothic", 11)
h3Bold = ("Century Gothic", 11, "bold")
h4 = ("Century Gothic", 9)

menuDict, storeOpenDict = excelConversion(wb)
# Main Frames
left_frame = Frame(root, width = 220, height = 150)
left_frame.grid(row = 0, column = 0, pady = 20, padx = 40)

left_frame2 = Frame(root, width = 220, height = 150)
#left_frame2.grid(row = 1, column = 0, pady = 23, padx = 30)

# Position Axis
w = 300
h = 200
x = 50
y = 100
root.geometry("%dx%d+%d+%d" % (w, h, x, y))

# Frame in Left Frame + Content
Label(left_frame, text = "Hello! Today is...", font = h1).place(x = 35, y = 20)
#Label(left_frame2, text = "Customisation", font = h1).place(x = 38, y = 20)

# Left frame content
now = datetime.datetime.now()
current = datetime.datetime.now().time()
currentdate = now.strftime("%d %b %Y")
currentime = now.strftime("%I:%M:%S %p")

Label(left_frame, text = currentdate, font = h3Italic).place(x = 60, y = 70)
clock = Label(left_frame, font = h3Italic)
clock.place(x = 68, y = 100)

# Right frame content
Label(text = "WELCOME TO NTU'S CANTEEN MASTER", font = h2, bg = "white").place(x = 300, y = 30)
Label(text = "I want to...", font = h3, bg = "white").place(x = 300, y = 80)

#today's stores button
load = Image.open('images/btn2.png')
render = ImageTk.PhotoImage(load)
img = Button(image = render, relief = FLAT, command = allstores, borderwidth = 0)
img.image = render
img.place(x = 300, y = 120)

#customise button
load = Image.open('images/btn12.png')
render = ImageTk.PhotoImage(load)
img = Button(image = render, relief = FLAT, command = custom, borderwidth = 0)
img.image = render
img.place(x = 450, y = 120)

#exit button
load = Image.open('images/exitbtn.png')
render = ImageTk.PhotoImage(load)
img = Button(image = render, relief = FLAT, command = root.destroy, borderwidth = 0)
img.image = render
img.place(x = 600, y = 120)

# About Canteen Master
Label(text = "ABOUT  US", font = h2, bg = "white").place(x = 430, y = 200)
load = Image.open('images/btn 4.png')
render = ImageTk.PhotoImage(load)
img = Label(image = render, relief = FLAT, bg = "white")
img.image = render
img.place(x = 280, y = 230)

tick()
root.mainloop() #display window until user closes it